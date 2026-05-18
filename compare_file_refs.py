#!/usr/bin/env python
"""Download the LegalSuite report (or use a CSV) and find missing Matter File Ref values."""

import argparse
import csv
import datetime as dt
import fnmatch
import ftplib
import json
import os
import re
import shutil
import sys
import warnings
from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path

from env_config import load_env_file

load_env_file()

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - runtime guard
    print("Missing dependency: openpyxl", file=sys.stderr)
    print("Install it with: python -m pip install openpyxl", file=sys.stderr)
    sys.exit(1)
else:
    warnings.filterwarnings(
        "ignore",
        message="Workbook contains no default style, apply openpyxl's default",
        module="openpyxl.styles.stylesheet",
    )

try:
    import requests
except ImportError:  # pragma: no cover - runtime guard
    print("Missing dependency: requests", file=sys.stderr)
    print("Install it with: python -m pip install requests", file=sys.stderr)
    sys.exit(1)

FTP_HOST = os.getenv("FTP_HOST", "")
FTP_USER = os.getenv("FTP_USER", "")
FTP_PASS = os.getenv("FTP_PASS", "")
FTP_TIMEOUT = 30
LEGALSUITE_API_BASE = "https://api.legalsuite.net"
LEGALSUITE_API_KEY = os.getenv("LEGALSUITE_API_KEY", "")
API_TIMEOUT = 60
LEGALSUITE_EMPLOYEE_ID = "1"
LEGALSUITE_ARCHIVE_STATUS = "2"
DEFAULT_CLIENT_IDS = [
    150307,
    334695,
    155128,
    334565,
    209250,
    334568,
    283850,
    334567,
    267742,
    334569,
]
DEFAULT_MATTER_TYPE_ID = 4
DEFAULT_ARCHIVE_STATUS = 0
LEGAL_SUITE_FIELDS = [
    "Matter.FileRef",
    "Matter.TheirRef",
    "Matter.ClientID",
]
LEGAL_SUITE_FILE_REF_FIELD = "Matter.FileRef"

TARGETS = [
    (
        "Standard Bank_ClaimsAmount",
        "Standard Bank Legal Claim Amount_Panel_L{year}_{month}_{day}_*.xlsx",
    ),
]

WEEKLY_TARGET = (
    "Standard Bank Claims Weekly Balancing",
    "Standard_Bank_Panel_L_Weekly_{year}{month}{day}.xlsx",
)

MATTER_REF_UPDATES_FTP_DIR = "Matter Ref Updates"


@dataclass(frozen=True)
class DateContext:
    date_str: str
    month_year: str
    year: str
    month: str
    day: str


@dataclass
class VerificationWorkbookState:
    source_path: str
    verification_path: str
    workbook: object
    header_indexes: dict[str, dict[str, int]]


class VerificationWorkbookRecorder:
    def __init__(self, verification_dir: Path, path_roots: list[Path]) -> None:
        self._verification_dir = verification_dir.resolve()
        self._path_roots = [path.resolve() for path in path_roots if path]
        self._states: dict[str, VerificationWorkbookState] = {}

    def record_row(
        self,
        source_path: Path,
        row_number: int,
        status: str,
        notes: str,
        get_response: object | None,
        verified_values: dict[str, object] | None = None,
        worksheet_name: str | None = None,
    ) -> Path:
        state = self._ensure_state(source_path)
        worksheet = self._resolve_worksheet(state, worksheet_name)
        values = {
            "Verification Status": status,
            "Verification Timestamp": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Verification Notes": notes,
            "Verification GET Response": self._serialize_response(get_response),
        }
        if verified_values:
            values.update(verified_values)

        for header_name, value in values.items():
            column_idx = self._ensure_column(state, worksheet.title, header_name)
            worksheet.cell(row=row_number, column=column_idx).value = value

        return state.verification_path

    def finalize(self) -> list[Path]:
        saved_paths: list[Path] = []
        for source_path, state in list(self._states.items()):
            state.workbook.save(state.verification_path)
            state.workbook.close()
            saved_paths.append(state.verification_path)
            del self._states[source_path]
        return sorted(saved_paths)

    def _ensure_state(self, source_path: Path) -> VerificationWorkbookState:
        source_abs = source_path.resolve()
        state = self._states.get(str(source_abs))
        if state is not None:
            return state

        verification_path = self._verification_path(source_abs)
        verification_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source_abs, verification_path)
        workbook = load_workbook(verification_path, read_only=False, data_only=False)
        state = VerificationWorkbookState(
            source_path=str(source_abs),
            verification_path=str(verification_path),
            workbook=workbook,
            header_indexes={},
        )
        self._states[str(source_abs)] = state
        return state

    def _verification_path(self, source_path: Path) -> Path:
        for root in self._path_roots:
            try:
                rel_path = source_path.relative_to(root)
            except ValueError:
                continue
            return self._verification_dir / rel_path
        return self._verification_dir / source_path.name

    @staticmethod
    def _normalize_header(value: object) -> str:
        if value is None:
            return ""
        return "".join(ch for ch in str(value).strip().lower() if ch.isalnum())

    @staticmethod
    def _serialize_response(response: object | None) -> str:
        if response in (None, ""):
            return ""
        text = json.dumps(response, default=str, ensure_ascii=True)
        if len(text) > 32000:
            return text[:31997] + "..."
        return text

    @staticmethod
    def _resolve_worksheet(state: VerificationWorkbookState, worksheet_name: str | None):
        if worksheet_name and worksheet_name in state.workbook.sheetnames:
            return state.workbook[worksheet_name]
        return state.workbook.active

    def _ensure_column(self, state: VerificationWorkbookState, worksheet_name: str, header_name: str) -> int:
        header_index = state.header_indexes.get(worksheet_name)
        worksheet = state.workbook[worksheet_name]
        if header_index is None:
            header_index = {}
            max_col = worksheet.max_column or 1
            for idx in range(1, max_col + 1):
                key = self._normalize_header(worksheet.cell(row=1, column=idx).value)
                if key and key not in header_index:
                    header_index[key] = idx
            state.header_indexes[worksheet_name] = header_index

        normalized_name = self._normalize_header(header_name)
        existing_idx = header_index.get(normalized_name)
        if existing_idx is not None:
            return existing_idx

        column_idx = (worksheet.max_column or 0) + 1
        worksheet.cell(row=1, column=column_idx).value = header_name
        header_index[normalized_name] = column_idx
        return column_idx


def build_date_context(today=None):
    if today is None:
        today = date.today()
    return DateContext(
        date_str=today.strftime("%Y-%m-%d"),
        month_year=today.strftime("%m-%Y"),
        year=today.strftime("%Y"),
        month=today.strftime("%m"),
        day=today.strftime("%d"),
    )


def build_week_bounds(today=None):
    if today is None:
        today = date.today()
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    return week_start, week_end


def build_legal_suite_headers(api_key):
    return {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/x-www-form-urlencoded",
    }


def get_matters_by_client_id(
    client_id,
    api_key,
    base_url,
    matter_type_id=DEFAULT_MATTER_TYPE_ID,
    archive_status=DEFAULT_ARCHIVE_STATUS,
    timeout=API_TIMEOUT,
):
    payload = [
        ("select[]", "Matter.FileRef"),
        ("select[]", "Matter.TheirRef"),
        ("select[]", "Matter.ClientID"),
        ("where[]", f"Matter.MatterTypeID,=,{matter_type_id}"),
        ("where[]", f"Matter.ClientID,=,{client_id}"),
        ("where[]", f"Matter.ArchiveStatus,=,{archive_status}"),
        ("archivestatusdescription", "Live"),
    ]

    response = requests.post(
        f"{base_url.rstrip('/')}/matter/get",
        headers=build_legal_suite_headers(api_key),
        data=payload,
        timeout=timeout,
    )
    response.raise_for_status()

    try:
        return response.json()
    except Exception:
        return response.text


def get_matters_by_client_ids(
    client_ids,
    api_key,
    base_url,
    matter_type_id=DEFAULT_MATTER_TYPE_ID,
    archive_status=DEFAULT_ARCHIVE_STATUS,
    timeout=API_TIMEOUT,
):
    combined_data = []

    for client_id in client_ids:
        result = get_matters_by_client_id(
            client_id=client_id,
            api_key=api_key,
            base_url=base_url,
            matter_type_id=matter_type_id,
            archive_status=archive_status,
            timeout=timeout,
        )

        if isinstance(result, dict) and "data" in result:
            combined_data.extend(result.get("data", []))
        else:
            print(f"Unexpected response for client {client_id}: {result}")

    return {"data": combined_data}


class LegalSuiteClient:
    def __init__(self, api_base, api_key, timeout=API_TIMEOUT):
        self._api_base = api_base.rstrip("/")
        self._api_key = api_key
        self._timeout = timeout

    def get_matter_by_fileref(self, file_ref):
        url = f"{self._api_base}/matter/get"
        data = {
            "where[]": f"Matter.FileRef,=,{file_ref}",
        }
        resp = requests.post(url, headers=self._headers(), data=data, timeout=self._timeout)
        resp.raise_for_status()
        payload = resp.json()
        items = get_legal_suite_records(payload)
        if not items:
            raise ValueError(f"No matter found for FileRef: {file_ref}")
        return items[0]

    def update_matter(self, payload):
        url = f"{self._api_base}/matter/update"
        resp = requests.post(url, headers=self._headers(), data=payload, timeout=self._timeout)
        resp.raise_for_status()
        try:
            return resp.json()
        except ValueError:
            return {"raw_response": resp.text}

    def _headers(self):
        return {
            "Authorization": f"Bearer {self._api_key}",
            "Content-Type": "application/x-www-form-urlencoded",
        }


def build_archive_payload(matter, logged_in_employee_id, archive_status=None):
    now = dt.datetime.now()
    payload = {}
    for key, value in matter.items():
        if isinstance(value, (dict, list, tuple, set)):
            continue
        payload[key] = value
    payload.update(
        {
            "loggedinemployeeid": str(logged_in_employee_id),
            "archiveflag": "1",
            "archivestatusdescription": "Archived",
            "formattedupdatedbydate": now.strftime("%d %b %Y"),
            "formattedupdatedbytime": now.strftime("%H:%M:%S"),
        }
    )
    if archive_status is not None:
        payload["archivestatus"] = str(archive_status)
    archive_no = (
        matter.get("archiveno")
        or matter.get("archive_no")
        or matter.get("archivenumber")
    )
    if archive_no is not None:
        payload["archiveno"] = str(archive_no)
    return {key: value for key, value in payload.items() if value not in ("", None)}


def build_pending_deletion_payload(matter, logged_in_employee_id):
    now = dt.datetime.now()
    payload = {}
    for key, value in matter.items():
        if isinstance(value, (dict, list, tuple, set)):
            continue
        payload[key] = value
    payload.update(
        {
            "loggedinemployeeid": str(logged_in_employee_id),
            "archiveflag": "0",
            "archivestatus": "1",
            "archivestatusdescription": "Pending Deletion",
            "formattedupdatedbydate": now.strftime("%d %b %Y"),
            "formattedupdatedbytime": now.strftime("%H:%M:%S"),
        }
    )
    archive_no = (
        matter.get("archiveno")
        or matter.get("archive_no")
        or matter.get("archivenumber")
    )
    if archive_no is not None:
        payload["archiveno"] = str(archive_no)
    return {key: value for key, value in payload.items() if value not in ("", None)}


def extract_update_error_text(result):
    if isinstance(result, dict):
        errors = result.get("errors")
        if errors:
            return str(errors)
        raw_response = result.get("raw_response")
        if raw_response:
            return str(raw_response)
    return str(result or "")


def is_archive_rejected_error(result):
    error_text = extract_update_error_text(result).lower()
    return (
        "you cannot archive a matter" in error_text
        or "you cannot archieve a matter" in error_text
    )


def normalize_compare_value(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def verify_matter_fields(expected_payload, fetched_matter, field_names, ignore_fields=None):
    ignored = set(ignore_fields or [])
    mismatches = []
    for field_name in field_names:
        if field_name in ignored:
            continue
        expected = normalize_compare_value(expected_payload.get(field_name))
        actual = normalize_compare_value(get_legal_suite_field(fetched_matter, field_name))
        if expected != actual:
            mismatches.append((field_name, expected, actual))
    return mismatches


def pattern_to_regex(pattern):
    escaped = re.escape(pattern)
    escaped = escaped.replace(r"\{year\}", r"(?P<year>\d{4})")
    escaped = escaped.replace(r"\{month\}", r"(?P<month>\d{2})")
    escaped = escaped.replace(r"\{day\}", r"(?P<day>\d{2})")
    escaped = escaped.replace(r"\*", r".*")
    return re.compile(f"^{escaped}$")


def list_ftp_dir(ftp, remote_dir):
    try:
        return ftp.nlst(remote_dir)
    except ftplib.error_perm:
        pass
    current_dir = None
    try:
        current_dir = ftp.pwd()
        ftp.cwd(remote_dir)
        listing = ftp.nlst()
        return [f"{remote_dir}/{name}" for name in listing]
    except ftplib.error_perm:
        return None
    finally:
        if current_dir:
            try:
                ftp.cwd(current_dir)
            except ftplib.all_errors:
                pass


def select_remote_file(remote_files, pattern):
    matches = []
    for path in remote_files:
        name = Path(path).name
        if fnmatch.fnmatch(name, pattern):
            matches.append(path)
    if not matches:
        return None
    matches.sort(key=lambda p: Path(p).name)
    return matches[-1]


def download_standard_file(
    output_dir,
    remote_dir,
    pattern,
    ftp_host,
    ftp_user,
    ftp_pass,
    ftp_timeout,
    target_date=None,
):
    ctx = build_date_context(today=target_date)
    formatted_pattern = pattern.format(
        year=ctx.year,
        month=ctx.month,
        day=ctx.day,
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    ftp = ftplib.FTP(ftp_host, timeout=ftp_timeout)
    try:
        ftp.login(ftp_user, ftp_pass)
        ftp.set_pasv(True)
        remote_files = list_ftp_dir(ftp, remote_dir)
        if not remote_files:
            raise SystemExit(f"No files found in FTP dir: {remote_dir}")
        selected = select_remote_file(remote_files, formatted_pattern)
        if not selected:
            raise SystemExit(
                f"No FTP file matched '{formatted_pattern}' in {remote_dir}"
            )
        remote_path = selected
        if "/" not in remote_path:
            remote_path = f"{remote_dir}/{remote_path}"
        local_path = output_dir / Path(selected).name
        with local_path.open("wb") as handle:
            ftp.retrbinary(f"RETR {remote_path}", handle.write)
        return local_path
    finally:
        try:
            ftp.quit()
        except ftplib.all_errors:
            ftp.close()


def download_weekly_file(
    output_dir,
    remote_dir,
    pattern,
    ftp_host,
    ftp_user,
    ftp_pass,
    ftp_timeout,
):
    week_start, week_end = build_week_bounds()
    regex = pattern_to_regex(pattern)
    output_dir.mkdir(parents=True, exist_ok=True)
    ftp = ftplib.FTP(ftp_host, timeout=ftp_timeout)
    try:
        ftp.login(ftp_user, ftp_pass)
        ftp.set_pasv(True)
        remote_files = list_ftp_dir(ftp, remote_dir)
        if not remote_files:
            raise SystemExit(f"No files found in FTP dir: {remote_dir}")
        matches = []
        for path in remote_files:
            name = Path(path).name
            match = regex.match(name)
            if not match:
                continue
            try:
                file_date = date(
                    int(match.group("year")),
                    int(match.group("month")),
                    int(match.group("day")),
                )
            except ValueError:
                continue
            if week_start <= file_date <= week_end:
                matches.append((file_date, path))
        if not matches:
            raise SystemExit(
                f"No weekly file matched {week_start} to {week_end} in {remote_dir}"
            )
        matches.sort(key=lambda item: (item[0], Path(item[1]).name))
        selected = matches[-1][1]
        remote_path = selected
        if "/" not in remote_path:
            remote_path = f"{remote_dir}/{remote_path}"
        local_path = output_dir / Path(selected).name
        with local_path.open("wb") as handle:
            ftp.retrbinary(f"RETR {remote_path}", handle.write)
        return local_path
    finally:
        try:
            ftp.quit()
        except ftplib.all_errors:
            ftp.close()


def upload_file_to_ftp_root(
    local_path,
    remote_dir,
    ftp_host,
    ftp_user,
    ftp_pass,
    ftp_timeout,
):
    ftp = ftplib.FTP(ftp_host, timeout=ftp_timeout)
    try:
        ftp.login(ftp_user, ftp_pass)
        ftp.set_pasv(True)
        ftp.cwd(remote_dir)
        remote_name = Path(local_path).name
        with Path(local_path).open("rb") as handle:
            ftp.storbinary(f"STOR {remote_name}", handle)
        return f"{remote_dir}/{remote_name}"
    finally:
        try:
            ftp.quit()
        except ftplib.all_errors:
            ftp.close()


def normalize(value, case_sensitive=False):
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    if not isinstance(value, str):
        value = str(value)
    value = value.strip()
    if not value:
        return None
    if value.isdigit():
        value = value.lstrip("0") or "0"
    return value if case_sensitive else value.lower()


def is_closed_status(value):
    if value is None:
        return False
    text = str(value).strip().lower()
    if not text:
        return False
    return text == "close" or text.startswith("closed")


def collect_closed_refs(found_rows, case_sensitive=False):
    closed_refs = []
    seen = set()
    for ref, status in found_rows:
        if not is_closed_status(status):
            continue
        norm = normalize(ref, case_sensitive=case_sensitive)
        if norm is None or norm in seen:
            continue
        seen.add(norm)
        closed_refs.append(ref)
    return closed_refs


def archive_closed_matters(
    file_refs,
    api_base,
    api_key,
    logged_in_employee_id,
    archive_status,
    timeout=API_TIMEOUT,
    dry_run=False,
    verbose=False,
):
    client = LegalSuiteClient(api_base, api_key, timeout=timeout)
    archived = 0
    failed = 0
    verification_fields = [
        "archiveflag",
        "archivestatus",
        "archivestatusdescription",
        "archiveno",
    ]
    for file_ref in file_refs:
        try:
            matter = client.get_matter_by_fileref(file_ref)
            payload = build_archive_payload(
                matter=matter,
                logged_in_employee_id=logged_in_employee_id,
                archive_status=archive_status,
            )
            if dry_run:
                print(f"Close dry-run: {file_ref}")
                if verbose:
                    print(json.dumps(payload, indent=2, default=str))
                archived += 1
                continue
            if verbose:
                print(f"Close request for {file_ref}:")
                print(json.dumps(payload, indent=2, default=str))
            try:
                result = client.update_matter(payload)
            except Exception as exc:
                print(f"Archive error for {file_ref}: {exc}; setting Pending Deletion...")
                pending_payload = build_pending_deletion_payload(
                    matter=matter,
                    logged_in_employee_id=logged_in_employee_id,
                )
                if verbose:
                    print(f"Pending Deletion request for {file_ref}:")
                    print(json.dumps(pending_payload, indent=2, default=str))
                pending_result = client.update_matter(pending_payload)
                if verbose:
                    print(f"Pending Deletion response for {file_ref}:")
                    print(json.dumps(pending_result, indent=2, default=str))
                archived_matter = client.get_matter_by_fileref(file_ref)
                mismatches = verify_matter_fields(
                    pending_payload,
                    archived_matter,
                    verification_fields,
                )
                if mismatches:
                    failed += 1
                    mismatch_text = ", ".join(
                        f"{field} expected={expected!r} actual={actual!r}"
                        for field, expected, actual in mismatches
                    )
                    print(
                        f"Pending Deletion verification failed for {file_ref}: {mismatch_text}",
                        file=sys.stderr,
                    )
                    continue
                archived += 1
                print(
                    f"Set to Pending Deletion: {file_ref} | "
                    f"archivestatusdescription={archived_matter.get('archivestatusdescription')}"
                )
                continue
            if verbose:
                print(f"Close response for {file_ref}:")
                print(json.dumps(result, indent=2, default=str))
            archived_matter = None
            if is_archive_rejected_error(result):
                print(f"Archive rejected for {file_ref}; setting Pending Deletion...")
                pending_payload = build_pending_deletion_payload(
                    matter=matter,
                    logged_in_employee_id=logged_in_employee_id,
                )
                if verbose:
                    print(f"Pending Deletion request for {file_ref}:")
                    print(json.dumps(pending_payload, indent=2, default=str))
                pending_result = client.update_matter(pending_payload)
                if verbose:
                    print(f"Pending Deletion response for {file_ref}:")
                    print(json.dumps(pending_result, indent=2, default=str))
                archived_matter = client.get_matter_by_fileref(file_ref)
                mismatches = verify_matter_fields(
                    pending_payload,
                    archived_matter,
                    verification_fields,
                )
                if mismatches:
                    failed += 1
                    mismatch_text = ", ".join(
                        f"{field} expected={expected!r} actual={actual!r}"
                        for field, expected, actual in mismatches
                    )
                    print(
                        f"Pending Deletion verification failed for {file_ref}: {mismatch_text}",
                        file=sys.stderr,
                    )
                    continue
                archived += 1
                print(
                    f"Set to Pending Deletion: {file_ref} | "
                    f"archivestatusdescription={archived_matter.get('archivestatusdescription')}"
                )
                continue
            archived_matter = client.get_matter_by_fileref(file_ref)
            archive_status_desc = str(
                archived_matter.get("archivestatusdescription")
                or archived_matter.get("ArchiveStatusDescription")
                or ""
            ).strip()
            if archive_status_desc.lower() == "live":
                print(f"Archive did not stick for {file_ref}; setting Pending Deletion...")
                pending_payload = build_pending_deletion_payload(
                    matter=archived_matter,
                    logged_in_employee_id=logged_in_employee_id,
                )
                if verbose:
                    print(f"Pending Deletion request for {file_ref}:")
                    print(json.dumps(pending_payload, indent=2, default=str))
                pending_result = client.update_matter(pending_payload)
                if verbose:
                    print(f"Pending Deletion response for {file_ref}:")
                    print(json.dumps(pending_result, indent=2, default=str))
                archived_matter = client.get_matter_by_fileref(file_ref)
                mismatches = verify_matter_fields(
                    pending_payload,
                    archived_matter,
                    verification_fields,
                )
                if mismatches:
                    failed += 1
                    mismatch_text = ", ".join(
                        f"{field} expected={expected!r} actual={actual!r}"
                        for field, expected, actual in mismatches
                    )
                    print(
                        f"Pending Deletion verification failed for {file_ref}: {mismatch_text}",
                        file=sys.stderr,
                    )
                    continue
                archived += 1
                print(
                    f"Set to Pending Deletion: {file_ref} | "
                    f"archivestatusdescription={archived_matter.get('archivestatusdescription')}"
                )
                continue
            mismatches = verify_matter_fields(
                payload,
                archived_matter,
                verification_fields,
                ignore_fields={"archiveno"},
            )
            if mismatches:
                failed += 1
                mismatch_text = ", ".join(
                    f"{field} expected={expected!r} actual={actual!r}"
                    for field, expected, actual in mismatches
                )
                print(
                    f"Archive verification failed for {file_ref}: {mismatch_text}",
                    file=sys.stderr,
                )
                continue
            archived += 1
            print(
                f"Closed matter: {file_ref} | "
                f"archivestatusdescription={archived_matter.get('archivestatusdescription')}"
            )
        except Exception as exc:
            failed += 1
            print(f"Close failed for {file_ref}: {exc}", file=sys.stderr)
    print(f"Close summary: archived={archived}, failed={failed}")


def find_column(headers, target_names):
    normalized_headers = []
    for name in headers:
        if name is None:
            normalized_headers.append(None)
        else:
            normalized_headers.append(str(name).strip().lower())
    for target in target_names:
        if not target:
            continue
        needle = target.strip().lower()
        for idx, name in enumerate(normalized_headers):
            if name == needle:
                return idx
    return None


def find_header_row(ws, target_names, max_rows=20):
    max_col = ws.max_column or 1
    for row_index in range(1, max_rows + 1):
        row = list(
            ws.iter_rows(
                min_row=row_index,
                max_row=row_index,
                max_col=max_col,
                values_only=True,
            )
        )
        if not row:
            continue
        headers = list(row[0])
        if find_column(headers, target_names) is not None:
            return row_index, headers
    return 1, list(
        next(
            ws.iter_rows(
                min_row=1,
                max_row=1,
                max_col=max_col,
                values_only=True,
            ),
            (),
        )
    )


def convert_csv_to_xlsx(csv_path, xlsx_path):
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Report")
    with csv_path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        for row in reader:
            ws.append(row)
    wb.save(xlsx_path)


def get_legal_suite_records(response):
    if isinstance(response, dict):
        for key in ("data", "records", "rows", "items", "result"):
            data = response.get(key)
            if isinstance(data, list):
                return data
        return []
    if isinstance(response, list):
        return response
    return []


def get_legal_suite_field(record, dotted_field):
    if isinstance(record, (list, tuple)):
        if dotted_field in LEGAL_SUITE_FIELDS:
            idx = LEGAL_SUITE_FIELDS.index(dotted_field)
            if idx < len(record):
                return record[idx]
        return None
    if not isinstance(record, dict):
        return None
    if dotted_field in record:
        return record.get(dotted_field)
    if "." not in dotted_field:
        return record.get(dotted_field)
    root, leaf = dotted_field.split(".", 1)
    nested = record.get(root)
    if isinstance(nested, dict):
        return nested.get(leaf)

    def normalize_key(value):
        return "".join(char for char in value.lower() if char.isalnum())

    candidates = {
        normalize_key(dotted_field),
        normalize_key(leaf),
        normalize_key(dotted_field.replace(".", "")),
    }
    for key, value in record.items():
        if normalize_key(str(key)) in candidates:
            return value
    return None


def extract_matters_file_refs(response, field_name=LEGAL_SUITE_FILE_REF_FIELD):
    values = []
    records = get_legal_suite_records(response)
    for record in records:
        raw = get_legal_suite_field(record, field_name)
        if raw is None:
            continue
        if not isinstance(raw, str):
            raw = str(raw)
        raw = raw.strip()
        if not raw:
            continue
        values.append(raw)
    return values


def write_legal_suite_report_xlsx(output_path, response):
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Report")
    ws.append(["Matter File Ref", "Matter Their Ref", "Matter Client ID"])
    records = get_legal_suite_records(response)
    for record in records:
        ws.append(
            [
                get_legal_suite_field(record, "Matter.FileRef"),
                get_legal_suite_field(record, "Matter.TheirRef"),
                get_legal_suite_field(record, "Matter.ClientID"),
            ]
        )
    wb.save(output_path)


def extract_csv_column(csv_path, column_name, case_sensitive=False):
    values = []
    with csv_path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        try:
            headers = next(reader)
        except StopIteration:
            raise ValueError(f"CSV file is empty: {csv_path}")
        idx = find_column(headers, [column_name])
        if idx is None:
            raise ValueError(
                f"Column '{column_name}' not found in CSV headers: {headers}"
            )
        for row in reader:
            if idx >= len(row):
                continue
            raw = row[idx]
            norm = normalize(raw, case_sensitive=case_sensitive)
            if norm is None:
                continue
            values.append(raw.strip())
    return values


def extract_excel_column(xlsx_path, column_name, aliases=None, case_sensitive=False):
    wb = load_workbook(xlsx_path, read_only=False, data_only=True)
    ws = wb.active
    target_names = [column_name] + (aliases or [])
    header_row, headers = find_header_row(ws, target_names)
    if not headers:
        raise ValueError(f"XLSX file is empty: {xlsx_path}")
    idx = find_column(headers, target_names)
    if idx is None:
        raise ValueError(
            f"Column '{column_name}' not found in XLSX headers: {headers}"
        )

    values = set()
    max_col = ws.max_column or 1
    for row in ws.iter_rows(
        min_row=header_row + 1,
        max_col=max_col,
        values_only=True,
    ):
        if idx >= len(row):
            continue
        norm = normalize(row[idx], case_sensitive=case_sensitive)
        if norm is None:
            continue
        values.add(norm)
    return values


def read_excel_column_values(xlsx_path, column_name, aliases=None):
    wb = load_workbook(xlsx_path, read_only=False, data_only=True)
    ws = wb.active
    target_names = [column_name] + (aliases or [])
    header_row, headers = find_header_row(ws, target_names)
    if not headers:
        raise ValueError(f"XLSX file is empty: {xlsx_path}")
    idx = find_column(headers, target_names)
    if idx is None:
        raise ValueError(
            f"Column '{column_name}' not found in XLSX headers: {headers}"
        )
    values = []
    max_col = ws.max_column or 1
    for row in ws.iter_rows(
        min_row=header_row + 1,
        max_col=max_col,
        values_only=True,
    ):
        if idx >= len(row):
            continue
        value = row[idx]
        if value is None:
            continue
        if isinstance(value, str):
            value = value.strip()
        if value == "":
            continue
        values.append(str(value).strip())
    return values


def build_weekly_status_map(
    weekly_path,
    file_ref_column,
    file_ref_aliases,
    status_aliases,
    case_sensitive=False,
):
    wb = load_workbook(weekly_path, read_only=False, data_only=True)
    ws = wb.active
    header_row, headers = find_header_row(ws, [file_ref_column] + status_aliases)
    if not headers:
        raise ValueError(f"XLSX file is empty: {weekly_path}")
    file_idx = find_column(headers, [file_ref_column] + (file_ref_aliases or []))
    if file_idx is None:
        raise ValueError(
            f"Column '{file_ref_column}' not found in XLSX headers: {headers}"
        )
    status_idx = find_column(headers, status_aliases)
    if status_idx is None:
        raise ValueError(
            f"Status column not found in XLSX headers: {headers}"
        )
    status_map = {}
    max_col = ws.max_column or 1
    for row in ws.iter_rows(
        min_row=header_row + 1,
        max_col=max_col,
        values_only=True,
    ):
        if file_idx >= len(row):
            continue
        file_ref = normalize(row[file_idx], case_sensitive=case_sensitive)
        if not file_ref:
            continue
        status_value = None
        if status_idx < len(row):
            status_value = row[status_idx]
        if isinstance(status_value, str):
            status_value = status_value.strip()
        if status_value is None:
            status_value = ""
        existing = status_map.get(file_ref, "")
        if not existing and status_value:
            status_map[file_ref] = status_value
        elif file_ref not in status_map:
            status_map[file_ref] = status_value
    return status_map


def build_weekly_lookup_maps(
    weekly_path,
    file_ref_column,
    matter_column,
    status_aliases,
    file_ref_aliases=None,
    matter_aliases=None,
    case_sensitive=False,
):
    wb = load_workbook(weekly_path, read_only=False, data_only=True)
    ws = wb.active
    target_names = [file_ref_column, matter_column] + status_aliases
    header_row, headers = find_header_row(ws, target_names)
    if not headers:
        raise ValueError(f"XLSX file is empty: {weekly_path}")

    file_idx = find_column(headers, [file_ref_column] + (file_ref_aliases or []))
    if file_idx is None:
        raise ValueError(
            f"Column '{file_ref_column}' not found in XLSX headers: {headers}"
        )
    matter_idx = find_column(headers, [matter_column] + (matter_aliases or []))
    if matter_idx is None:
        raise ValueError(
            f"Column '{matter_column}' not found in XLSX headers: {headers}"
        )
    status_idx = find_column(headers, status_aliases)
    if status_idx is None:
        raise ValueError(f"Status column not found in XLSX headers: {headers}")

    by_file_ref = {}
    by_matter = {}
    max_col = ws.max_column or 1
    for row in ws.iter_rows(
        min_row=header_row + 1,
        max_col=max_col,
        values_only=True,
    ):
        raw_file_ref = row[file_idx] if file_idx < len(row) else None
        raw_matter = row[matter_idx] if matter_idx < len(row) else None
        raw_status = row[status_idx] if status_idx < len(row) else None
        file_ref = normalize(raw_file_ref, case_sensitive=case_sensitive)
        matter = normalize(raw_matter, case_sensitive=case_sensitive)
        status = ""
        if raw_status is not None:
            status = str(raw_status).strip()
        entry = {
            "file_ref": "" if raw_file_ref is None else str(raw_file_ref).strip(),
            "matter": "" if raw_matter is None else str(raw_matter).strip(),
            "status": status,
        }
        if file_ref:
            existing = by_file_ref.get(file_ref)
            if existing is None or (not existing.get("status") and status):
                by_file_ref[file_ref] = entry
        if matter:
            existing = by_matter.get(matter)
            if existing is None or (not existing.get("status") and status):
                by_matter[matter] = entry
    return by_file_ref, by_matter


def write_missing_to_xlsx(output_path, header, values):
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Missing Matter File Ref")
    ws.append([header])
    for value in values:
        ws.append([value])
    wb.save(output_path)


def write_found_to_xlsx(output_path, header, rows):
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Matter Found From Weekly")
    ws.append([header, "Status"])
    for ref, status in rows:
        ws.append([ref, status])
    wb.save(output_path)


def write_not_found_to_xlsx(output_path, header, values):
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Matter Not Found From Weekly")
    ws.append([header])
    for value in values:
        ws.append([value])
    wb.save(output_path)


def write_matter_ref_updates_xlsx(output_path, rows):
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Sheet1")
    ws.append(["Matter File Ref", "Their Ref"])
    for row in rows:
        ws.append([row["matter_file_ref"], row["their_ref"]])
    wb.save(output_path)


def record_missing_verification(
    recorder,
    workbook_path,
    missing_refs,
    standard_path,
    standard_column,
):
    for row_number, ref in enumerate(missing_refs, start=2):
        recorder.record_row(
            workbook_path,
            row_number,
            status="Verified Missing",
            notes=f"Not found in {standard_path.name} column '{standard_column}'.",
            get_response={
                "file_ref": ref,
                "standard_file": str(standard_path),
                "standard_column": standard_column,
                "match_found": False,
            },
            verified_values={
                "Verified Against File": standard_path.name,
                "Verified Against Column": standard_column,
            },
        )


def record_weekly_found_verification(recorder, workbook_path, found_rows, weekly_path):
    for row_number, (ref, status_value) in enumerate(found_rows, start=2):
        status_text = str(status_value or "").strip()
        note = "Found in weekly report."
        if not status_text:
            note = "Found in weekly report but Current Status was blank."
        recorder.record_row(
            workbook_path,
            row_number,
            status="Verified Weekly Match",
            notes=note,
            get_response={
                "file_ref": ref,
                "weekly_file": str(weekly_path),
                "status": status_text,
                "match_found": True,
            },
            verified_values={
                "Verified Weekly File": weekly_path.name,
                "Verified Current Status": status_text,
            },
        )


def record_weekly_not_found_verification(recorder, workbook_path, not_found_refs, weekly_path):
    for row_number, ref in enumerate(not_found_refs, start=2):
        recorder.record_row(
            workbook_path,
            row_number,
            status="Verified Weekly Missing",
            notes=f"Not found in weekly file {weekly_path.name}.",
            get_response={
                "file_ref": ref,
                "weekly_file": str(weekly_path),
                "match_found": False,
            },
            verified_values={
                "Verified Weekly File": weekly_path.name,
            },
        )


def record_matter_ref_updates_verification(recorder, workbook_path, rows, weekly_path):
    for row_number, row in enumerate(rows, start=2):
        recorder.record_row(
            workbook_path,
            row_number,
            status="Verified Weekly Matter Match",
            notes=(
                "Weekly record matched by Their Ref in the Matter column; "
                "Matter Ref update required."
            ),
            get_response={
                "matter_file_ref": row["matter_file_ref"],
                "their_ref": row["their_ref"],
                "weekly_file_ref": row["weekly_file_ref"],
                "weekly_status": row["status"],
                "weekly_file": str(weekly_path),
                "match_found": True,
                "lookup_method": "Their Ref -> Weekly Matter",
            },
            verified_values={
                "Verified Weekly File": weekly_path.name,
                "Verified Their Ref": row["their_ref"],
                "Verified Weekly File Ref": row["weekly_file_ref"],
                "Verified Current Status": row["status"],
            },
        )


def write_verification_summary(verification_dir, summary):
    verification_dir.mkdir(parents=True, exist_ok=True)
    json_path = verification_dir / "verification_summary.json"
    txt_path = verification_dir / "verification_summary.txt"
    json_path.write_text(json.dumps(summary, ensure_ascii=True, indent=2), encoding="utf-8")

    lines = [
        "compare_file_refs verification summary",
        f"Run timestamp: {summary['run_timestamp']}",
        f"Report source: {summary['report_source']}",
        f"Converted report: {summary['converted_report']}",
        f"Standard claim file: {summary['standard_file']}",
        f"Compared report refs: {summary['report_ref_count']}",
        f"Matched refs: {summary['matched_count']}",
        f"Missing refs: {summary['missing_count']}",
    ]
    if summary.get("weekly_file"):
        lines.append(f"Weekly file: {summary['weekly_file']}")
        lines.append(f"Weekly found refs: {summary['weekly_found_count']}")
        lines.append(f"Weekly not found refs: {summary['weekly_not_found_count']}")
        lines.append(f"Matter Ref update rows: {summary['matter_ref_updates_count']}")
    if summary.get("matter_ref_updates_remote_path"):
        lines.append(
            f"Matter Ref updates FTP path: {summary['matter_ref_updates_remote_path']}"
        )
    if summary.get("verification_workbooks"):
        lines.append("Verification workbooks:")
        for path in summary["verification_workbooks"]:
            lines.append(f"  - {path}")
    txt_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return json_path, txt_path


def parse_args():
    parser = argparse.ArgumentParser(
        description=(
            "Download the LegalSuite report (or use a CSV), then compare "
            "Matter File Ref against File Ref in the Standard Bank file."
        )
    )
    parser.add_argument(
        "--csv",
        type=Path,
        help="Path to Report (SUP).CSV (skips LegalSuite API download)",
    )
    parser.add_argument(
        "--csv-column",
        default="Matter File Ref",
        help="Column name in the CSV to compare",
    )
    parser.add_argument(
        "--api-key",
        default=os.getenv("LEGALSUITE_API_KEY", LEGALSUITE_API_KEY),
        help="LegalSuite API key (or set LEGALSUITE_API_KEY)",
    )
    parser.add_argument(
        "--api-base-url",
        default=LEGALSUITE_API_BASE,
        help="LegalSuite API base URL",
    )
    parser.add_argument(
        "--api-timeout",
        default=API_TIMEOUT,
        type=int,
        help="LegalSuite API timeout in seconds",
    )
    parser.add_argument(
        "--debug-legal-suite-output",
        type=Path,
        help="Write raw LegalSuite API response to JSON for debugging",
    )
    parser.add_argument(
        "--client-ids",
        type=int,
        nargs="+",
        default=DEFAULT_CLIENT_IDS,
        help="LegalSuite client IDs for the report",
    )
    parser.add_argument(
        "--matter-type-id",
        type=int,
        default=DEFAULT_MATTER_TYPE_ID,
        help="LegalSuite MatterTypeID filter",
    )
    parser.add_argument(
        "--archive-status",
        type=int,
        default=DEFAULT_ARCHIVE_STATUS,
        help="LegalSuite ArchiveStatus filter",
    )
    parser.add_argument(
        "--close-closed",
        action="store_true",
        help="Close matters in LegalSuite when weekly status is Closed",
    )
    parser.add_argument(
        "--close-dry-run",
        action="store_true",
        help="Fetch matters but do not update LegalSuite",
    )
    parser.add_argument(
        "--close-verbose",
        action="store_true",
        help="Print LegalSuite close payloads and responses",
    )
    parser.add_argument(
        "--logged-in-employee-id",
        default=LEGALSUITE_EMPLOYEE_ID,
        help="LegalSuite logged-in employee ID for updates",
    )
    parser.add_argument(
        "--close-archive-status",
        default=LEGALSUITE_ARCHIVE_STATUS,
        help="Archive status to send when closing matters",
    )
    parser.add_argument(
        "--standard",
        type=Path,
        help="Path to Standard Bank Legal Claim Amount file (skips FTP download)",
    )
    parser.add_argument(
        "--ftp-host",
        default=FTP_HOST,
        help="FTP host for Standard Bank files",
    )
    parser.add_argument(
        "--ftp-user",
        default=FTP_USER,
        help="FTP user for Standard Bank files",
    )
    parser.add_argument(
        "--ftp-pass",
        default=FTP_PASS,
        help="FTP password for Standard Bank files",
    )
    parser.add_argument(
        "--ftp-timeout",
        default=FTP_TIMEOUT,
        type=int,
        help="FTP timeout in seconds",
    )
    parser.add_argument(
        "--ftp-dir",
        default=TARGETS[0][0],
        help="FTP directory for Standard Bank files",
    )
    parser.add_argument(
        "--ftp-pattern",
        default=TARGETS[0][1],
        help="FTP filename pattern for today's Standard Bank file",
    )
    parser.add_argument(
        "--day",
        type=int,
        default=0,
        help="Days back for Standard Bank file date (e.g. 1 for yesterday)",
    )
    parser.add_argument(
        "--weekly-ftp-pattern",
        default=WEEKLY_TARGET[1],
        help="FTP filename pattern for today's weekly balancing file",
    )
    parser.add_argument(
        "--converted",
        default="Report (SUP).xlsx",
        type=Path,
        help="Output XLSX path for LegalSuite report or converted CSV",
    )
    parser.add_argument(
        "--output",
        default="Missing_Matter_File_Ref.xlsx",
        type=Path,
        help="Output XLSX with missing Matter File Ref values",
    )
    parser.add_argument(
        "--weekly-found-output",
        default="Matter_Found_From_Weekly.xlsx",
        type=Path,
        help="Output XLSX with missing refs found in weekly report",
    )
    parser.add_argument(
        "--weekly-not-found-output",
        default="Matter_Not_Found_From_Weekly.xlsx",
        type=Path,
        help="Output XLSX with missing refs not found in weekly report",
    )
    parser.add_argument(
        "--matter-ref-updates-output",
        default="Panel L Matter Ref Updates.xlsx",
        type=Path,
        help="Output XLSX for Weekly Matter Ref update rows",
    )
    parser.add_argument(
        "--verification-dir",
        default=Path("verification"),
        type=Path,
        help="Directory for verification workbooks and summaries",
    )
    parser.add_argument(
        "--standard-column",
        default="File Ref",
        help="Column name in the Standard Bank file",
    )
    parser.add_argument(
        "--case-sensitive",
        action="store_true",
        help="Use case-sensitive comparisons",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    verification_dir = args.verification_dir
    recorder = VerificationWorkbookRecorder(
        verification_dir=verification_dir,
        path_roots=[Path.cwd()],
    )
    weekly_path = None
    found = []
    not_found = []
    matter_ref_updates = []
    verification_summary = {
        "run_timestamp": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "report_source": "",
        "converted_report": "",
        "standard_file": "",
        "weekly_file": "",
        "report_ref_count": 0,
        "matched_count": 0,
        "missing_count": 0,
        "weekly_found_count": 0,
        "weekly_not_found_count": 0,
        "matter_ref_updates_count": 0,
        "missing_output": str(args.output.resolve()),
        "weekly_found_output": str(args.weekly_found_output.resolve()),
        "weekly_not_found_output": str(args.weekly_not_found_output.resolve()),
        "matter_ref_updates_output": str(args.matter_ref_updates_output.resolve()),
        "matter_ref_updates_remote_path": "",
        "verification_workbooks": [],
    }

    if args.csv:
        if not args.csv.exists():
            raise SystemExit(f"CSV not found: {args.csv}")
        report_refs = extract_csv_column(
            args.csv,
            args.csv_column,
            case_sensitive=args.case_sensitive,
        )
        convert_csv_to_xlsx(args.csv, args.converted)
        print(f"Converted CSV -> {args.converted}")
        verification_summary["report_source"] = str(args.csv.resolve())
    else:
        if not args.api_key:
            raise SystemExit(
                "LegalSuite API key missing. Use --api-key or set LEGALSUITE_API_KEY."
            )
        response = get_matters_by_client_ids(
            client_ids=args.client_ids,
            api_key=args.api_key,
            base_url=args.api_base_url,
            matter_type_id=args.matter_type_id,
            archive_status=args.archive_status,
            timeout=args.api_timeout,
        )
        if args.debug_legal_suite_output:
            with args.debug_legal_suite_output.open("w", encoding="utf-8") as handle:
                json.dump(response, handle, ensure_ascii=True, indent=2)
            print(f"Wrote LegalSuite debug output -> {args.debug_legal_suite_output}")
        records = get_legal_suite_records(response)
        report_refs = extract_matters_file_refs(response)
        write_legal_suite_report_xlsx(args.converted, response)
        print(f"Wrote LegalSuite report -> {args.converted}")
        verification_summary["report_source"] = "LegalSuite API"
        if not records or not report_refs:
            print(
                "LegalSuite response parsed but no Matter File Ref values were found. "
                "Use --debug-legal-suite-output to inspect the raw response."
            )

    if not report_refs:
        raise SystemExit("No Matter File Ref values found in the report data.")
    verification_summary["converted_report"] = str(args.converted.resolve())
    verification_summary["report_ref_count"] = len(report_refs)

    if args.standard is None:
        target_date = date.today() - timedelta(days=args.day)
        args.standard = download_standard_file(
            output_dir=Path.cwd(),
            remote_dir=args.ftp_dir,
            pattern=args.ftp_pattern,
            ftp_host=args.ftp_host,
            ftp_user=args.ftp_user,
            ftp_pass=args.ftp_pass,
            ftp_timeout=args.ftp_timeout,
            target_date=target_date,
        )
        print(f"Downloaded Standard Bank file -> {args.standard}")
    elif not args.standard.exists():
        raise SystemExit(f"Standard Bank XLSX not found: {args.standard}")
    verification_summary["standard_file"] = str(args.standard.resolve())

    matter_refs = report_refs
    file_refs = extract_excel_column(
        args.standard,
        args.standard_column,
        aliases=[
            "Matter",
            "Matter File Ref",
            "Matter File",
            "File Reference",
        ],
        case_sensitive=args.case_sensitive,
    )

    seen = set()
    missing = []
    for raw in matter_refs:
        norm = normalize(raw, case_sensitive=args.case_sensitive)
        if norm is None:
            continue
        if norm in file_refs or norm in seen:
            continue
        seen.add(norm)
        missing.append(raw.strip())

    write_missing_to_xlsx(args.output, "Matter File Ref", missing)
    record_missing_verification(
        recorder,
        args.output,
        missing,
        args.standard,
        args.standard_column,
    )

    print(f"Missing count: {len(missing)}")
    print(f"Wrote missing refs -> {args.output}")
    verification_summary["missing_count"] = len(missing)
    verification_summary["matched_count"] = max(len(matter_refs) - len(missing), 0)

    if missing:
        weekly_path = download_weekly_file(
            output_dir=Path.cwd(),
            remote_dir=WEEKLY_TARGET[0],
            pattern=args.weekly_ftp_pattern,
            ftp_host=args.ftp_host,
            ftp_user=args.ftp_user,
            ftp_pass=args.ftp_pass,
            ftp_timeout=args.ftp_timeout,
        )
        print(f"Downloaded weekly report -> {weekly_path}")
        verification_summary["weekly_file"] = str(weekly_path.resolve())

        try:
            missing_refs = read_excel_column_values(
                args.output,
                "Matter File Ref",
                aliases=[args.csv_column],
            )
            weekly_statuses = build_weekly_status_map(
                weekly_path,
                args.standard_column,
                file_ref_aliases=[
                    "Matter",
                    "Matter File Ref",
                    "Matter File",
                    "File Reference",
                ],
                status_aliases=[
                    "Status",
                    "Matter Status",
                    "Current Status",
                    "Claim Status",
                    "Stage",
                ],
                case_sensitive=args.case_sensitive,
            )
            _, weekly_matter_map = build_weekly_lookup_maps(
                weekly_path,
                args.standard_column,
                "Matter",
                status_aliases=[
                    "Status",
                    "Matter Status",
                    "Current Status",
                    "Claim Status",
                    "Stage",
                ],
                file_ref_aliases=[
                    "Matter",
                    "Matter File Ref",
                    "Matter File",
                    "File Reference",
                ],
                matter_aliases=["Account Number", "Their Ref"],
                case_sensitive=args.case_sensitive,
            )
        except ValueError as exc:
            print(f"Weekly lookup skipped: {exc}")
            verification_summary["weekly_found_count"] = 0
            verification_summary["weekly_not_found_count"] = len(missing)
            verification_paths = [str(path) for path in recorder.finalize()]
            verification_summary["verification_workbooks"] = verification_paths
            json_path, txt_path = write_verification_summary(verification_dir, verification_summary)
            print(f"Wrote verification summary -> {json_path}")
            print(f"Wrote verification notes -> {txt_path}")
            return

        for ref in missing_refs:
            norm = normalize(ref, case_sensitive=args.case_sensitive)
            status = weekly_statuses.get(norm)
            if status is None:
                not_found.append(ref)
                continue
            if status == "":
                print(f"{ref}: status not provided")
                found.append((ref, ""))
            else:
                print(f"{ref}: {status}")
                found.append((ref, status))
        unresolved_after_weekly = []
        if not_found:
            if not args.api_key:
                print(
                    "Weekly Matter fallback skipped: missing LegalSuite API key "
                    "(use --api-key or LEGALSUITE_API_KEY)."
                )
                unresolved_after_weekly = list(not_found)
            else:
                client = LegalSuiteClient(
                    api_base=args.api_base_url,
                    api_key=args.api_key,
                    timeout=args.api_timeout,
                )
                for ref in not_found:
                    try:
                        matter = client.get_matter_by_fileref(ref)
                    except Exception as exc:
                        print(f"{ref}: LegalSuite lookup failed: {exc}")
                        unresolved_after_weekly.append(ref)
                        continue
                    their_ref = get_legal_suite_field(matter, "Matter.TheirRef")
                    their_ref_text = ""
                    if their_ref is not None:
                        their_ref_text = str(their_ref).strip()
                    if not their_ref_text:
                        print(f"{ref}: LegalSuite Their Ref not available")
                        unresolved_after_weekly.append(ref)
                        continue
                    weekly_match = weekly_matter_map.get(
                        normalize(their_ref_text, case_sensitive=args.case_sensitive)
                    )
                    if not weekly_match:
                        print(
                            f"{ref}: not found in weekly Matter column using Their Ref {their_ref_text}"
                        )
                        unresolved_after_weekly.append(ref)
                        continue
                    status = str(weekly_match.get("status") or "").strip()
                    wrong_weekly_ref = str(weekly_match.get("file_ref") or "").strip()
                    print(
                        f"{ref}: found in weekly Matter column using Their Ref {their_ref_text} "
                        f"| weekly File Ref={wrong_weekly_ref or '<blank>'} "
                        f"| status={status or 'status not provided'}"
                    )
                    matter_ref_updates.append(
                        {
                            "matter_file_ref": ref,
                            "their_ref": their_ref_text,
                            "weekly_file_ref": wrong_weekly_ref,
                            "status": status,
                        }
                    )
                    found.append((ref, status))
                not_found = unresolved_after_weekly
        closed_refs = collect_closed_refs(found, case_sensitive=args.case_sensitive)
        if closed_refs:
            if not args.close_closed:
                print(
                    "Closed matters detected in weekly report. "
                    "Use --close-closed to update LegalSuite."
                )
            elif not args.api_key:
                print(
                    "Close skipped: missing LegalSuite API key (use --api-key or LEGALSUITE_API_KEY)."
                )
            else:
                archive_closed_matters(
                    closed_refs,
                    api_base=args.api_base_url,
                    api_key=args.api_key,
                    logged_in_employee_id=args.logged_in_employee_id,
                    archive_status=args.close_archive_status,
                    timeout=args.api_timeout,
                    dry_run=args.close_dry_run,
                    verbose=args.close_verbose,
                )
        if not_found:
            print("Not found in weekly report:")
            for ref in not_found:
                print(ref)
        if found:
            write_found_to_xlsx(
                args.weekly_found_output,
                "Matter File Ref",
                found,
            )
            print(f"Wrote weekly found refs -> {args.weekly_found_output}")
            record_weekly_found_verification(
                recorder,
                args.weekly_found_output,
                found,
                weekly_path,
            )
        if matter_ref_updates:
            write_matter_ref_updates_xlsx(
                args.matter_ref_updates_output,
                matter_ref_updates,
            )
            print(f"Wrote matter ref updates -> {args.matter_ref_updates_output}")
            remote_upload_path = upload_file_to_ftp_root(
                local_path=args.matter_ref_updates_output,
                remote_dir=MATTER_REF_UPDATES_FTP_DIR,
                ftp_host=args.ftp_host,
                ftp_user=args.ftp_user,
                ftp_pass=args.ftp_pass,
                ftp_timeout=args.ftp_timeout,
            )
            verification_summary["matter_ref_updates_remote_path"] = remote_upload_path
            print(f"Uploaded matter ref updates -> {remote_upload_path}")
            record_matter_ref_updates_verification(
                recorder,
                args.matter_ref_updates_output,
                matter_ref_updates,
                weekly_path,
            )
        if not_found:
            write_not_found_to_xlsx(
                args.weekly_not_found_output,
                "Matter File Ref",
                not_found,
            )
            print(f"Wrote weekly not found refs -> {args.weekly_not_found_output}")
            record_weekly_not_found_verification(
                recorder,
                args.weekly_not_found_output,
                not_found,
                weekly_path,
            )

    verification_summary["weekly_found_count"] = len(found)
    verification_summary["weekly_not_found_count"] = len(not_found)
    verification_summary["matter_ref_updates_count"] = len(matter_ref_updates)
    verification_paths = [str(path) for path in recorder.finalize()]
    verification_summary["verification_workbooks"] = verification_paths
    json_path, txt_path = write_verification_summary(verification_dir, verification_summary)
    print(f"Wrote verification summary -> {json_path}")
    print(f"Wrote verification notes -> {txt_path}")


if __name__ == "__main__":
    main()
